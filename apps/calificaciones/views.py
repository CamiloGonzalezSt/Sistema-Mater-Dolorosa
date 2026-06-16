from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.views.generic import ListView, TemplateView, View

from apps.alumnos.models import Matricula
from apps.asistencia.views import ROLES_DOCENTES, cursos_asignatura_de
from apps.core.mixins import RoleRequiredMixin

from .forms import CalificacionForm, FiltroNotasExportForm
from .models import Calificacion, Evaluacion, PeriodoEvaluacion


def evaluaciones_de(user):
    """Evaluaciones que el usuario puede operar: las de sus asignaturas, o todas si es admin."""
    return (
        Evaluacion.objects.filter(curso_asignatura__in=cursos_asignatura_de(user))
        .select_related(
            'curso_asignatura__curso__nivel',
            'curso_asignatura__asignatura',
            'periodo',
            'tipo',
        )
        .order_by('-fecha')
    )


class SeleccionEvaluacionView(RoleRequiredMixin, ListView):
    """La tabla de evaluaciones solo aparece con los 3 filtros completos
    (curso + profesor + período). El profesor queda fijado a sí mismo."""

    allowed_roles = ROLES_DOCENTES
    template_name = 'calificaciones/seleccionar.html'
    context_object_name = 'evaluaciones'

    def get_filtro(self):
        from django.contrib.auth import get_user_model

        from apps.academico.models import Curso

        from .forms import FiltroEvaluacionesForm

        form = FiltroEvaluacionesForm(self.request.GET or None)
        user = self.request.user
        if user.role != 'admin':
            # Propiedad: solo él mismo y sus cursos
            form.fields['profesor'].queryset = get_user_model().objects.filter(pk=user.pk)
            form.fields['profesor'].initial = user.pk
            form.fields['curso'].queryset = Curso.objects.filter(
                curso_asignaturas__profesor=user
            ).select_related('nivel').distinct()
        return form

    def get_queryset(self):
        self.form_filtro = self.get_filtro()
        if not self.form_filtro.is_bound or not self.form_filtro.is_valid():
            return evaluaciones_de(self.request.user).none()
        datos = self.form_filtro.cleaned_data
        return evaluaciones_de(self.request.user).filter(
            curso_asignatura__curso=datos['curso'],
            curso_asignatura__profesor=datos['profesor'],
            periodo=datos['periodo'],
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_filtro'] = self.form_filtro
        context['filtros_aplicados'] = (
            self.form_filtro.is_bound and self.form_filtro.is_valid()
        )
        return context


class IngresarNotasView(RoleRequiredMixin, TemplateView):
    allowed_roles = ROLES_DOCENTES
    template_name = 'calificaciones/ingresar.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role not in self.allowed_roles:
            return super().dispatch(request, *args, **kwargs)  # mixin resuelve login/403
        # Propiedad: 404 si la evaluación no es de una asignatura del profesor
        self.evaluacion = get_object_or_404(
            evaluaciones_de(request.user), pk=kwargs['pk']
        )
        return super().dispatch(request, *args, **kwargs)

    def get_forms(self, data=None):
        matriculas = (
            Matricula.objects.filter(
                curso=self.evaluacion.curso_asignatura.curso,
                estado__in=[Matricula.Estado.REGULAR, Matricula.Estado.REPITENTE],
            )
            .select_related('alumno__usuario')
            .order_by('alumno__usuario__last_name', 'alumno__usuario__first_name')
        )
        existentes = {
            c.matricula_id: c
            for c in Calificacion.objects.filter(
                evaluacion=self.evaluacion, matricula__in=matriculas
            )
        }
        forms = []
        for m in matriculas:
            instancia = existentes.get(m.pk) or Calificacion(
                evaluacion=self.evaluacion, matricula=m
            )
            # Filas nuevas dejadas en blanco son válidas y se omiten al guardar:
            # no se obliga a calificar a todo el curso de una vez.
            forms.append(
                (m, CalificacionForm(
                    data,
                    instance=instancia,
                    prefix=f'm{m.pk}',
                    empty_permitted=instancia.pk is None,
                    use_required_attribute=False,
                ))
            )
        return forms

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault('filas', self.get_forms())
        context['evaluacion'] = self.evaluacion
        return context

    def post(self, request, *args, **kwargs):
        filas = self.get_forms(request.POST)
        if all(form.is_valid() for _, form in filas):
            guardadas = 0
            with transaction.atomic():
                for _, form in filas:
                    # Fila nueva sin datos → no se crea calificación
                    if form.instance.pk is None and not form.has_changed():
                        continue
                    form.save()  # el middleware de simple-history registra al usuario
                    guardadas += 1
            messages.success(
                request, f'Notas guardadas: {guardadas} de {len(filas)} alumnos.'
            )
            return redirect('calificaciones:ingresar', pk=self.evaluacion.pk)
        context = self.get_context_data(filas=filas)
        return self.render_to_response(context)


class ExportarNotasExcelView(RoleRequiredMixin, View):
    """Exporta a Excel todas las notas de un curso+período.
    Una fila por alumno, una columna por evaluación + columna de promedio final."""

    allowed_roles = ROLES_DOCENTES

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        form = FiltroNotasExportForm(request.GET or None)
        if not form.is_valid():
            return HttpResponse('Parámetros inválidos. Vuelve y selecciona curso y período.', status=400)

        curso = form.cleaned_data['curso']
        periodo = form.cleaned_data['periodo']

        # Solo las CursoAsignatura del curso accesibles por el usuario
        cas = cursos_asignatura_de(request.user).filter(
            curso=curso, anio_escolar=periodo.anio_escolar
        ).select_related('asignatura').order_by('asignatura__nombre')

        evaluaciones = (
            Evaluacion.objects.filter(curso_asignatura__in=cas, periodo=periodo)
            .select_related('curso_asignatura__asignatura', 'tipo')
            .order_by('curso_asignatura__asignatura__nombre', 'fecha')
        )

        matriculas = (
            Matricula.objects.filter(
                curso=curso,
                estado__in=[Matricula.Estado.REGULAR, Matricula.Estado.REPITENTE],
            )
            .select_related('alumno__usuario')
            .order_by('alumno__usuario__last_name', 'alumno__usuario__first_name')
        )

        # Índice de calificaciones: (matricula_id, evaluacion_id) → nota
        cals = Calificacion.objects.filter(
            evaluacion__in=evaluaciones, matricula__in=matriculas
        ).values_list('matricula_id', 'evaluacion_id', 'nota')
        notas_idx = {(m, e): float(n) for m, e, n in cals}

        wb = openpyxl.Workbook()
        ws = wb.active
        # Título máx 31 chars (límite Excel)
        ws.title = f'{curso.nivel.nombre[:10]}{curso.letra} {periodo.anio_escolar}'[:31]

        navy = PatternFill('solid', fgColor='010D61')
        font_header = Font(bold=True, color='FFFFFF')

        ev_list = list(evaluaciones)

        # Fila 1: nombre asignatura (agrupada)
        ws.cell(1, 1, 'Alumno').font = font_header
        ws.cell(1, 1).fill = navy
        ws.cell(1, 2, 'RUT').font = font_header
        ws.cell(1, 2).fill = navy
        col = 3
        for ev in ev_list:
            c = ws.cell(1, col, ev.curso_asignatura.asignatura.nombre)
            c.font = font_header
            c.fill = navy
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            col += 1
        c = ws.cell(1, col, 'Promedio')
        c.font = font_header
        c.fill = navy
        c.alignment = Alignment(horizontal='center')

        # Fila 2: nombre evaluación
        ws.cell(2, 1, '').fill = navy
        ws.cell(2, 2, '').fill = navy
        col = 3
        for ev in ev_list:
            c = ws.cell(2, col, f'{ev.nombre} ({ev.tipo.nombre})')
            c.font = Font(italic=True, color='FFFFFF')
            c.fill = navy
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            col += 1
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 40

        # Filas de alumnos
        for fila_idx, mat in enumerate(matriculas, start=3):
            nombre = mat.alumno.usuario.get_full_name() or mat.alumno.rut_alumno
            ws.cell(fila_idx, 1, nombre)
            ws.cell(fila_idx, 2, mat.alumno.rut_alumno)
            notas_fila = []
            col = 3
            for ev in ev_list:
                nota = notas_idx.get((mat.pk, ev.pk))
                c = ws.cell(fila_idx, col, nota)
                if nota is not None:
                    c.number_format = '0.0'
                    notas_fila.append(nota)
                col += 1
            promedio = round(sum(notas_fila) / len(notas_fila), 1) if notas_fila else None
            c = ws.cell(fila_idx, col, promedio)
            if promedio is not None:
                c.number_format = '0.0'
                c.font = Font(bold=True)

        # Anchos de columna
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 13
        for i in range(3, 3 + len(ev_list) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

        import unicodedata
        def slug(s):
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s.replace(' ', '_')

        nombre_archivo = (
            f"notas_{slug(curso.nivel.nombre)}{curso.letra}"
            f"_{periodo.anio_escolar}_{slug(periodo.nombre)}.xlsx"
        )
        respuesta = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        respuesta['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(respuesta)
        return respuesta


class MisNotasView(RoleRequiredMixin, TemplateView):
    """Vista de notas para el alumno autenticado.
    Muestra sus calificaciones agrupadas por asignatura y período."""

    allowed_roles = ['alumno']
    template_name = 'calificaciones/mis_notas.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = getattr(self.request.user, 'perfil_alumno', None)
        if not perfil:
            context['sin_perfil'] = True
            return context

        anio = timezone.localdate().year
        matricula = (
            Matricula.objects.filter(alumno=perfil, anio_escolar=anio)
            .select_related('curso__nivel', 'curso__profesor_jefe__usuario')
            .first()
        )
        context['matricula'] = matricula
        if not matricula:
            return context

        # Todas las calificaciones del alumno en el año en curso
        cals = (
            Calificacion.objects.filter(matricula=matricula)
            .select_related(
                'evaluacion__curso_asignatura__asignatura',
                'evaluacion__periodo',
                'evaluacion__tipo',
            )
            .order_by(
                'evaluacion__curso_asignatura__asignatura__nombre',
                'evaluacion__periodo__fecha_inicio',
                'evaluacion__fecha',
            )
        )

        # Agrupar: { periodo: { asignatura: [cal, ...] } }
        from collections import defaultdict
        agrupado = defaultdict(lambda: defaultdict(list))
        for cal in cals:
            periodo = cal.evaluacion.periodo
            asignatura = cal.evaluacion.curso_asignatura.asignatura
            agrupado[periodo][asignatura].append(cal)

        # Calcular promedios y armar estructura ordenada
        resumen = []
        for periodo in sorted(agrupado.keys(), key=lambda p: p.fecha_inicio):
            asignaturas = []
            for asignatura, califs in sorted(
                agrupado[periodo].items(), key=lambda x: x[0].nombre
            ):
                notas = [float(c.nota) for c in califs]
                promedio = round(sum(notas) / len(notas), 1) if notas else None
                asignaturas.append({
                    'asignatura': asignatura,
                    'calificaciones': califs,
                    'promedio': promedio,
                })
            resumen.append({'periodo': periodo, 'asignaturas': asignaturas})

        context['resumen'] = resumen
        return context
