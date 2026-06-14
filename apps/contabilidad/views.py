from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.text import slugify
from django.views.generic import CreateView, FormView, ListView, TemplateView, View

from apps.alumnos.models import Matricula
from apps.core.mixins import RoleRequiredMixin

from .forms import GeneracionCobrosForm, PagoForm
from .models import Cobro, Pago
from .reportes import comprobante_pago_pdf


class ListadoCobrosView(RoleRequiredMixin, ListView):
    """Gestión de cobros: exclusiva del administrador (matriz RBAC)."""

    allowed_roles = ['admin']
    template_name = 'contabilidad/cobros.html'
    context_object_name = 'cobros'
    paginate_by = 50

    def get_queryset(self):
        qs = Cobro.objects.select_related(
            'matricula__alumno__usuario', 'matricula__curso__nivel', 'tipo_arancel'
        )
        estado = self.request.GET.get('estado', '')
        if estado in Cobro.Estado.values:
            qs = qs.filter(estado=estado)
        periodo = self.request.GET.get('periodo', '').strip()
        if periodo:
            qs = qs.filter(periodo=periodo)
        busqueda = self.request.GET.get('q', '').strip()
        if busqueda:
            qs = qs.filter(
                Q(matricula__alumno__usuario__first_name__icontains=busqueda)
                | Q(matricula__alumno__usuario__last_name__icontains=busqueda)
                | Q(matricula__alumno__rut_alumno__icontains=busqueda)
            )
        curso_id = self.request.GET.get('curso', '')
        if curso_id.isdigit():
            qs = qs.filter(matricula__curso_id=curso_id)
        return qs

    def get_context_data(self, **kwargs):
        from apps.academico.models import Curso

        context = super().get_context_data(**kwargs)
        context['estados'] = Cobro.Estado.choices
        context['filtro_estado'] = self.request.GET.get('estado', '')
        context['filtro_periodo'] = self.request.GET.get('periodo', '')
        context['filtro_curso'] = self.request.GET.get('curso', '')
        context['busqueda'] = self.request.GET.get('q', '')
        context['cursos'] = Curso.objects.filter(
            matriculas__isnull=False
        ).select_related('nivel').distinct()
        pendientes = Cobro.objects.exclude(
            estado__in=[Cobro.Estado.PAGADO, Cobro.Estado.CONDONADO]
        )
        context['total_por_cobrar'] = (
            pendientes.aggregate(t=Sum('monto'))['t'] or 0
        ) - (
            Pago.objects.filter(cobro__in=pendientes).aggregate(
                t=Sum('monto_pagado'))['t'] or 0
        )
        return context


class ExportarCobrosExcelView(ListadoCobrosView):
    """Exporta a Excel el listado de cobros con los filtros aplicados."""

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cobros'
        encabezados = ['Alumno', 'RUT', 'Curso', 'Arancel', 'Período',
                       'Monto', 'Pagado', 'Saldo', 'Estado', 'Vencimiento']
        ws.append(encabezados)
        relleno = PatternFill('solid', fgColor='010D61')
        for celda in ws[1]:
            celda.font = Font(bold=True, color='FFFFFF')
            celda.fill = relleno

        for cobro in self.get_queryset():
            alumno = cobro.matricula.alumno
            ws.append([
                alumno.usuario.get_full_name() or alumno.rut_alumno,
                alumno.rut_alumno,
                str(cobro.matricula.curso),
                cobro.tipo_arancel.nombre,
                cobro.periodo,
                float(cobro.monto),
                float(cobro.total_pagado),
                float(cobro.saldo_pendiente),
                cobro.get_estado_display(),
                cobro.fecha_vencimiento.strftime('%d-%m-%Y') if cobro.fecha_vencimiento else '',
            ])

        for i, ancho in enumerate([28, 13, 18, 18, 12, 12, 12, 12, 12, 13], start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        respuesta = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        respuesta['Content-Disposition'] = 'attachment; filename="cobros.xlsx"'
        wb.save(respuesta)
        return respuesta


class GenerarCobrosView(RoleRequiredMixin, FormView):
    allowed_roles = ['admin']
    template_name = 'contabilidad/generar.html'
    form_class = GeneracionCobrosForm
    success_url = reverse_lazy('contabilidad:cobros')

    def form_valid(self, form):
        datos = form.cleaned_data
        matriculas = Matricula.objects.filter(
            estado__in=[Matricula.Estado.REGULAR, Matricula.Estado.REPITENTE]
        )
        if datos['curso']:
            matriculas = matriculas.filter(curso=datos['curso'])
        existentes = set(
            Cobro.objects.filter(
                tipo_arancel=datos['tipo_arancel'], periodo=datos['periodo'],
                matricula__in=matriculas,
            ).values_list('matricula_id', flat=True)
        )
        monto = datos['monto'] or datos['tipo_arancel'].monto_base
        creados = 0
        with transaction.atomic():
            for matricula in matriculas:
                if matricula.pk in existentes:
                    continue
                Cobro.objects.create(
                    matricula=matricula, tipo_arancel=datos['tipo_arancel'],
                    periodo=datos['periodo'], monto=monto,
                    fecha_vencimiento=datos['fecha_vencimiento'])
                creados += 1
        messages.success(
            self.request,
            f'Cobros generados: {creados}. Omitidos por ya existir: {len(existentes)}.',
        )
        return super().form_valid(form)


class RegistrarPagoView(RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    form_class = PagoForm
    template_name = 'contabilidad/pago_form.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role in self.allowed_roles:
            self.cobro = get_object_or_404(
                Cobro.objects.select_related(
                    'matricula__alumno__usuario', 'tipo_arancel'),
                pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cobro'] = self.cobro
        return context

    def form_valid(self, form):
        form.instance.cobro = self.cobro
        form.instance.registrado_por = self.request.user
        respuesta = super().form_valid(form)
        messages.success(
            self.request,
            f'Pago registrado. Estado del cobro: '
            f'{self.cobro.get_estado_display()}.',
        )
        return respuesta

    def get_success_url(self):
        return reverse('contabilidad:cobros')


class ComprobantePagoPDFView(RoleRequiredMixin, View):
    """Descarga del comprobante: admin, o el apoderado vinculado al alumno."""

    allowed_roles = ['admin', 'apoderado']

    def get(self, request, *args, **kwargs):
        pago = get_object_or_404(
            Pago.objects.select_related(
                'cobro__matricula__alumno__usuario',
                'cobro__tipo_arancel', 'registrado_por'),
            pk=kwargs['pk'])
        if request.user.role == 'apoderado':
            es_su_pupilo = pago.cobro.matricula.alumno.vinculos_apoderados.filter(
                apoderado__usuario=request.user
            ).exists()
            if not es_su_pupilo:
                raise Http404
        pdf = comprobante_pago_pdf(pago)
        usuario_alumno = pago.cobro.matricula.alumno.usuario
        nombre_archivo = 'comprobante_{}_{}_{}.pdf'.format(
            slugify(usuario_alumno.first_name) or 'alumno',
            slugify(usuario_alumno.last_name) or 'sin-apellido',
            slugify(pago.cobro.tipo_arancel.nombre),
        )
        respuesta = HttpResponse(pdf, content_type='application/pdf')
        respuesta['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return respuesta


class MiCuentaView(RoleRequiredMixin, TemplateView):
    """Estado de cuenta del apoderado: solo lectura, solo sus pupilos."""

    allowed_roles = ['apoderado']
    template_name = 'contabilidad/mi_cuenta.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        perfil = getattr(self.request.user, 'perfil_apoderado', None)
        pupilos = []
        if perfil:
            for vinculo in perfil.vinculos_alumnos.select_related('alumno__usuario'):
                alumno = vinculo.alumno
                cobros = (
                    Cobro.objects.filter(matricula__alumno=alumno)
                    .select_related('tipo_arancel', 'matricula__curso__nivel')
                    .prefetch_related('pagos')
                    .order_by('-fecha_vencimiento')
                )
                saldo_total = sum(
                    (c.saldo_pendiente for c in cobros
                     if c.estado != Cobro.Estado.CONDONADO),
                    start=0,
                )
                pupilos.append(
                    {'alumno': alumno, 'cobros': cobros, 'saldo_total': saldo_total}
                )
        context['pupilos'] = pupilos
        return context
